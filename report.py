import os
import io
import math
import certifi
os.environ["SSL_CERT_FILE"] = certifi.where()
os.environ["REQUESTS_CA_BUNDLE"] = os.environ["SSL_CERT_FILE"]

import pandas as pd
import numpy as np
import yfinance as yf
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage

# ===== 參數 =====
INPUT_FILE  = "tickers.xlsx"        # 讀取股票代碼
OUTPUT_FILE = "ETF100_report.xlsx"  # 報告輸出
PERIOD      = "6mo"                 # 期間：6mo / 1y / 2y...
FIGSIZE     = (12, 9)
DPI         = 150

# ====== 工具函式（不用 pandas-ta，自行實作各指標） ======
def ema(series: pd.Series, span: int):
    return series.ewm(span=span, adjust=False).mean()

def rma(series: pd.Series, length: int):
    """Wilder's RMA（常用於 RSI、DMI）"""
    return series.ewm(alpha=1/length, adjust=False).mean()

def macd_series(close: pd.Series, fast=12, slow=26, signal=9):
    ema_fast = ema(close, fast)
    ema_slow = ema(close, slow)
    macd = ema_fast - ema_slow
    signal_line = ema(macd, signal)
    hist = macd - signal_line
    return macd, signal_line, hist

def rsi_series(close: pd.Series, length=14):
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = (-delta).clip(lower=0)
    avg_gain = rma(gain, length)
    avg_loss = rma(loss, length)
    rs = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    return rsi

def stoch_kdj(high: pd.Series, low: pd.Series, close: pd.Series, k_len=14, d_len=3, smooth_k=3):
    ll = low.rolling(k_len, min_periods=k_len).min()
    hh = high.rolling(k_len, min_periods=k_len).max()
    raw_k = 100 * (close - ll) / (hh - ll)
    k = raw_k.rolling(smooth_k, min_periods=smooth_k).mean()
    d = k.rolling(d_len, min_periods=d_len).mean()
    j = 3 * k - 2 * d
    return k, d, j

def dmi_adx(high: pd.Series, low: pd.Series, close: pd.Series, length=14):
    # True Range
    prev_close = close.shift(1)
    tr = pd.concat([
        high - low,
        (high - prev_close).abs(),
        (low - prev_close).abs()
    ], axis=1).max(axis=1)

    # +DM, -DM
    up_move = high.diff()
    down_move = -low.diff()
    plus_dm = np.where((up_move > down_move) & (up_move > 0), up_move, 0.0)
    minus_dm = np.where((down_move > up_move) & (down_move > 0), down_move, 0.0)
    plus_dm = pd.Series(plus_dm, index=high.index)
    minus_dm = pd.Series(minus_dm, index=high.index)

    # 平滑（Wilder RMA）
    tr_rma = rma(tr, length)
    plus_dm_rma = rma(plus_dm, length)
    minus_dm_rma = rma(minus_dm, length)

    plus_di = 100 * (plus_dm_rma / tr_rma.replace(0, np.nan))
    minus_di = 100 * (minus_dm_rma / tr_rma.replace(0, np.nan))
    dx = 100 * ( (plus_di - minus_di).abs() / (plus_di + minus_di).replace(0, np.nan) )
    adx = rma(dx, length)
    return plus_di, minus_di, adx

def vwap_series(high: pd.Series, low: pd.Series, close: pd.Series, volume: pd.Series):
    typical = (high + low + close) / 3.0
    cum_tp_vol = (typical * volume).cumsum()
    cum_vol = volume.cumsum()
    return cum_tp_vol / cum_vol.replace(0, np.nan)

def obv_series(close: pd.Series, volume: pd.Series):
    sign = np.sign(close.diff().fillna(0))
    # 常見實作：漲加量、跌減量、平盤 0（你也可改成沿用前值）
    return (sign * volume).fillna(0).cumsum()

def psy_series(close: pd.Series, length=12):
    up = (close.diff() > 0).astype(int)
    return up.rolling(length, min_periods=length).mean() * 100

def williams_r(high: pd.Series, low: pd.Series, close: pd.Series, length=14):
    hh = high.rolling(length, min_periods=length).max()
    ll = low.rolling(length, min_periods=length).min()
    return -100 * (hh - close) / (hh - ll)

def bias_series(close: pd.Series, length=6):
    ma = close.rolling(length, min_periods=length).mean()
    return (close - ma) / ma * 100

# ====== 資料抓取與處理 ======
def safe_download(ticker: str, period: str = PERIOD) -> pd.DataFrame:
    try:
        df = yf.download(
            ticker,
            period=period,
            auto_adjust=False,
            progress=False,
            threads=False,
            group_by="column",   # <<< 關鍵：欄位在最外層
        )
        if df is None or df.empty:
            return pd.DataFrame()

        # 聰明展平：找出哪一層有 OHLC 欄位
        if isinstance(df.columns, pd.MultiIndex):
            fields = {"open","high","low","close","adj close","volume"}
            lvl0 = [str(x).strip().lower() for x in df.columns.get_level_values(0)]
            lvl1 = [str(x).strip().lower() for x in df.columns.get_level_values(1)]
            if set(lvl0) & fields:
                # 欄位在第 0 層
                df.columns = df.columns.get_level_values(0)
            elif set(lvl1) & fields:
                # 欄位在第 1 層
                df.columns = df.columns.get_level_values(1)
            else:
                # 萬一兩層都不像，就逐一映射
                new_cols = []
                for a, b in df.columns:
                    a_l = str(a).strip().lower()
                    b_l = str(b).strip().lower()
                    if a_l in fields:
                        new_cols.append(a)
                    elif b_l in fields:
                        new_cols.append(b)
                    else:
                        new_cols.append(str(a))
                df.columns = new_cols
                print(f"   flattened columns: {list(df.columns)}")

        return df.dropna(how="all")
    except Exception:
        return pd.DataFrame()


def add_indicators(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        print("⚠️ raw is empty in add_indicators")
        return df

    out = normalize_ohlcv_columns(df.copy())

    # ---- Debug：看看到底有什麼欄位 ----
    print("   columns:", list(out.columns))
    print("   tail:\n", out.tail(2))

    # 選擇基準價格
    if "Adj Close" in out.columns:
        price = out["Adj Close"]
        price_col_name = "Adj Close"
    elif "Close" in out.columns:
        price = out["Close"]
        price_col_name = "Close"
    elif "Price" in out.columns:
        price = out["Price"]
        price_col_name = "Price"
    else:
        # 沒有任何 close 類欄位→嘗試自建
        if all(c in out.columns for c in ["Open", "High", "Low"]):
            price = (out["Open"] + out["High"] + out["Low"]) / 3.0
            out["Close"] = price
            price_col_name = "Close*synthetic"
            print("   ✅ 自建 Close（(O+H+L)/3）")
        else:
            print("   ❌ 仍找不到可用價欄，放棄此檔")
            return pd.DataFrame()

    # 確保必要欄位存在
    for col in ["Open", "High", "Low", "Close", "Volume"]:
        if col not in out.columns:
            if col == "Volume":
                out[col] = 0
            elif col == "Close":
                out[col] = price
            else:
                out[col] = price  # 沒開高低→用 price 近似
    print(f"   ✔️ 使用價欄：{price_col_name}")

    # ===== 指標（與你前版一致）=====
    out["MA3"]  = price.rolling(3,  min_periods=3).mean()
    out["MA5"]  = price.rolling(5,  min_periods=5).mean()
    out["MA10"] = price.rolling(10, min_periods=10).mean()

    # VWAP
    typical = (out["High"] + out["Low"] + out["Close"]) / 3.0
    cum_tp_vol = (typical * out["Volume"]).cumsum()
    cum_vol = out["Volume"].cumsum().replace(0, np.nan)
    out["VWAP"] = cum_tp_vol / cum_vol

    # MACD
    def ema(s, span): return s.ewm(span=span, adjust=False).mean()
    ema_fast = ema(price, 12)
    ema_slow = ema(price, 26)
    out["MACD"]   = ema_fast - ema_slow
    out["SIGNAL"] = ema(out["MACD"], 9)
    out["HIST"]   = out["MACD"] - out["SIGNAL"]

    # DMI/ADX
    prev_close = out["Close"].shift(1)
    tr = pd.concat([
        out["High"] - out["Low"],
        (out["High"] - prev_close).abs(),
        (out["Low"] - prev_close).abs()
    ], axis=1).max(axis=1)
    up_move = out["High"].diff()
    down_move = -out["Low"].diff()
    plus_dm  = pd.Series(np.where((up_move > down_move) & (up_move > 0), up_move, 0.0), index=out.index)
    minus_dm = pd.Series(np.where((down_move > up_move) & (down_move > 0), down_move, 0.0), index=out.index)
    def rma(s, n): return s.ewm(alpha=1/n, adjust=False).mean()
    tr_rma = rma(tr, 14)
    pdm_rma = rma(plus_dm, 14)
    mdm_rma = rma(minus_dm, 14)
    out["+DI"] = 100 * (pdm_rma / tr_rma.replace(0, np.nan))
    out["-DI"] = 100 * (mdm_rma / tr_rma.replace(0, np.nan))
    dx = 100 * (out["+DI"] - out["-DI"]).abs() / (out["+DI"] + out["-DI"]).replace(0, np.nan)
    out["ADX"] = rma(dx, 14)

    # RSI
    delta = price.diff()
    gain  = delta.clip(lower=0)
    loss  = (-delta).clip(lower=0)
    avg_gain = rma(gain, 14)
    avg_loss = rma(loss, 14)
    rs = avg_gain / avg_loss.replace(0, np.nan)
    out["RSI"] = 100 - (100 / (1 + rs))

    # KD & J
    ll = out["Low"].rolling(14, min_periods=14).min()
    hh = out["High"].rolling(14, min_periods=14).max()
    raw_k = 100 * (out["Close"] - ll) / (hh - ll)
    k = raw_k.rolling(3, min_periods=3).mean()
    d = k.rolling(3, min_periods=3).mean()
    out["%K"] = k
    out["%D"] = d
    out["%J"] = 3 * k - 2 * d

    # 其他
    sign = np.sign(out["Close"].diff().fillna(0))
    out["OBV"] = (sign * out["Volume"]).fillna(0).cumsum()
    up = (out["Close"].diff() > 0).astype(int)
    out["PSY"] = up.rolling(12, min_periods=12).mean() * 100
    hh2 = out["High"].rolling(14, min_periods=14).max()
    ll2 = out["Low"].rolling(14, min_periods=14).min()
    out["W%R"] = -100 * (hh2 - out["Close"]) / (hh2 - ll2)
    ma6 = price.rolling(6, min_periods=6).mean()
    out["BIAS6"] = (price - ma6) / ma6 * 100

    return out


def normalize_ohlcv_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    if isinstance(df.columns, pd.MultiIndex):
        # 經過 safe_download 應該已處理，不過再保險一次
        df.columns = df.columns.get_level_values(0)

    mapping = {
        "open": "Open",
        "high": "High",
        "low": "Low",
        "close": "Close",
        "adjclose": "Adj Close",
        "adj close": "Adj Close",
        "adjusted close": "Adj Close",
        "volume": "Volume",
        "price": "Close",
    }
    new_cols = []
    for c in df.columns:
        key = str(c).strip().lower().replace("_", " ").replace("-", " ")
        new_cols.append(mapping.get(key, c))
    df.columns = new_cols
    return df



# ====== 繪圖 ======
def pick_price_column(df: pd.DataFrame) -> str | None:
    for c in ["Close", "Adj Close", "Price"]:
        if c in df.columns:
            return c
    # 再做一次模糊找
    for c in df.columns:
        if "close" in str(c).lower():
            return c
    return None

def plot_six_panel(ticker: str, df: pd.DataFrame) -> bytes:
    if df is None or df.empty:
        raise ValueError("empty dataframe")
    price_col = pick_price_column(df)
    if price_col is None:
        raise ValueError("no price column to plot")

    fig, axes = plt.subplots(6, 1, sharex=True, figsize=FIGSIZE)
    ax1, ax2, ax3, ax4, ax5, ax6 = axes

    ax1.set_title(f"{ticker} - Price/MA/VWAP", loc="left")
    ax1.plot(df.index, df[price_col], label=price_col)
    for col in ["MA3", "MA5", "MA10", "VWAP"]:
        if col in df.columns:
            ax1.plot(df.index, df[col], label=col)
    ax1.grid(True, alpha=0.3)
    ax1.legend(loc="upper left", ncol=5, fontsize=8)
    ax1b = ax1.twinx()
    if "Volume" in df.columns:
        ax1b.bar(df.index, df["Volume"], alpha=0.25, width=1.0, label="Volume")
        ax1b.set_ylabel("Volume")

    # 其餘子圖不變…
    # （你原本的 MACD / DMI / RSI / KD / OBV-PSY-W%R-BIAS6 繪圖邏輯照舊）
    # ... 略 ...


    price_col = pick_price_column(df)
    if price_col is None:
        raise ValueError("no price column to plot")

    fig, axes = plt.subplots(6, 1, sharex=True, figsize=FIGSIZE)
    ax1, ax2, ax3, ax4, ax5, ax6 = axes

    # 1) Price + MAs + VWAP + Volume
    ax1.set_title(f"{ticker} - Price/MA/VWAP", loc="left")
    ax1.plot(df.index, df[price_col], label=price_col)
    for col in ["MA3", "MA5", "MA10", "VWAP"]:
        if col in df.columns:
            ax1.plot(df.index, df[col], label=col)
    ax1.grid(True, alpha=0.3)
    ax1.legend(loc="upper left", ncol=5, fontsize=8)
    ax1b = ax1.twinx()
    if "Volume" in df.columns:
        ax1b.bar(df.index, df["Volume"], alpha=0.25, width=1.0, label="Volume")
        ax1b.set_ylabel("Volume")

    # 2) MACD
    ax2.set_title("MACD", loc="left")
    if {"MACD", "SIGNAL", "HIST"}.issubset(df.columns):
        ax2.plot(df.index, df["MACD"], label="MACD")
        ax2.plot(df.index, df["SIGNAL"], label="Signal")
        ax2.bar(df.index, df["HIST"], width=1.0, alpha=0.4, label="Hist")
    ax2.grid(True, alpha=0.3)
    ax2.legend(loc="upper left", ncol=3, fontsize=8)

    # 3) DMI
    ax3.set_title("DMI", loc="left")
    for col in ["+DI", "-DI", "ADX"]:
        if col in df.columns:
            ax3.plot(df.index, df[col], label=col)
    ax3.grid(True, alpha=0.3)
    ax3.legend(loc="upper left", ncol=3, fontsize=8)

    # 4) RSI
    ax4.set_title("RSI(14)", loc="left")
    if "RSI" in df.columns:
        ax4.plot(df.index, df["RSI"], label="RSI")
        ax4.axhline(70, ls="--", lw=1)
        ax4.axhline(30, ls="--", lw=1)
    ax4.grid(True, alpha=0.3)
    ax4.legend(loc="upper left", fontsize=8)

    # 5) KD
    ax5.set_title("KD", loc="left")
    for col in ["%K", "%D", "%J"]:
        if col in df.columns:
            ax5.plot(df.index, df[col], label=col)
    ax5.grid(True, alpha=0.3)
    ax5.legend(loc="upper left", ncol=3, fontsize=8)

    # 6) OBV / PSY / W%R / BIAS6
    ax6.set_title("OBV / PSY / W%R / BIAS6", loc="left")
    for col in ["OBV", "PSY", "W%R", "BIAS6"]:
        if col in df.columns:
            ax6.plot(df.index, df[col], label=col)
    ax6.grid(True, alpha=0.3)
    ax6.legend(loc="upper left", ncol=4, fontsize=8)

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=DPI, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


# ====== Excel 輸出 ======
def ensure_workbook(path: str) -> Workbook:
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        ws = wb["Sheet"]
        ws.title = "README"
        ws["A1"] = f"Generated at {datetime.now()}"
    return wb

def write_report(ticker: str, df: pd.DataFrame, img_bytes: bytes, wb: Workbook):
    # 圖表分頁
    if ticker in wb.sheetnames:
        ws_chart = wb[ticker]
        for row in ws_chart["A1:Z200"]:
            for cell in row:
                cell.value = None
    else:
        ws_chart = wb.create_sheet(title=ticker)
    img = XLImage(io.BytesIO(img_bytes))
    img.anchor = "A1"
    ws_chart.add_image(img)

    # 數據分頁
    data_name = f"{ticker}_data"
    if data_name in wb.sheetnames:
        wb.remove(wb[data_name])
    ws_data = wb.create_sheet(title=data_name)
    cols = ["Date"] + list(df.columns)
    ws_data.append(cols)
    for _, row in df.reset_index().iterrows():
        values = [row["Date"]]
        for c in df.columns:
            v = row[c]
            if pd.isna(v):
                values.append(None)
            else:
                values.append(float(v) if isinstance(v, (np.floating, float, np.integer, int)) else v)
        ws_data.append(values)

def main():
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(f"找不到 {INPUT_FILE}，請建立含 'Ticker' 欄位的 Excel。")
    codes = pd.read_excel(INPUT_FILE)["Ticker"].dropna().astype(str).tolist()
    if not codes:
        raise ValueError("在 tickers.xlsx 中沒有讀到任何 Ticker。")

    wb = ensure_workbook(OUTPUT_FILE)

    for t in codes:
        print(f"[{t}] downloading...")
        raw = safe_download(t, PERIOD)
        if raw.empty:
            print("  -> 無資料（download），略過")
            continue

        df = add_indicators(raw)
        if df.empty:
            print("  -> 無資料（indicators），略過")
            # 額外印一下原始欄位，幫你排查
            print("   (debug) 原始欄位：", list(raw.columns))
            print("   (debug) 原始 tail:\n", raw.tail(2))
            continue

        try:
            img = plot_six_panel(t, df)
        except Exception as e:
            print(f"  -> 繪圖失敗：{e}，略過")
            continue

        write_report(t, df, img, wb)
        print("  -> 完成")

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 已輸出：{OUTPUT_FILE}")

if __name__ == "__main__":
    main()

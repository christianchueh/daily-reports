# app.py
import os
import io
import math
import certifi
os.environ["SSL_CERT_FILE"] = certifi.where()
os.environ["REQUESTS_CA_BUNDLE"] = os.environ["SSL_CERT_FILE"]

import datetime as dt
import pandas as pd
import numpy as np
import yfinance as yf
import matplotlib.pyplot as plt
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# ---------- 指標工具（不用 pandas-ta） ----------
def ema(series: pd.Series, span: int):
    return series.ewm(span=span, adjust=False).mean()

def rma(series: pd.Series, length: int):
    return series.ewm(alpha=1/length, adjust=False).mean()

def macd_series(close: pd.Series, fast=12, slow=26, signal=9):
    ema_fast = ema(close, fast)
    ema_slow = ema(close, slow)
    macd = ema_fast - ema_slow
    signal_line = ema(macd, signal)
    hist = macd - signal_line
    return macd, signal_line, hist

def rsi_series(close: pd.Series, length=14):
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = (-delta).clip(lower=0)
    avg_gain = rma(gain, length)
    avg_loss = rma(loss, length)
    rs = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    return rsi

def stoch_kdj(high: pd.Series, low: pd.Series, close: pd.Series, k_len=14, d_len=3, smooth_k=3):
    ll = low.rolling(k_len, min_periods=k_len).min()
    hh = high.rolling(k_len, min_periods=k_len).max()
    raw_k = 100 * (close - ll) / (hh - ll)
    k = raw_k.rolling(smooth_k, min_periods=smooth_k).mean()
    d = k.rolling(d_len, min_periods=d_len).mean()
    j = 3 * k - 2 * d
    return k, d, j

def dmi_adx(high: pd.Series, low: pd.Series, close: pd.Series, length=14):
    prev_close = close.shift(1)
    tr = pd.concat([
        high - low,
        (high - prev_close).abs(),
        (low - prev_close).abs()
    ], axis=1).max(axis=1)

    up_move = high.diff()
    down_move = -low.diff()
    plus_dm = pd.Series(np.where((up_move > down_move) & (up_move > 0), up_move, 0.0), index=high.index)
    minus_dm = pd.Series(np.where((down_move > up_move) & (down_move > 0), down_move, 0.0), index=high.index)

    tr_rma = rma(tr, length)
    plus_dm_rma = rma(plus_dm, length)
    minus_dm_rma = rma(minus_dm, length)

    plus_di = 100 * (plus_dm_rma / tr_rma.replace(0, np.nan))
    minus_di = 100 * (minus_dm_rma / tr_rma.replace(0, np.nan))
    dx = 100 * ((plus_di - minus_di).abs() / (plus_di + minus_di).replace(0, np.nan))
    adx = rma(dx, length)
    return plus_di, minus_di, adx

def vwap_series(high: pd.Series, low: pd.Series, close: pd.Series, volume: pd.Series):
    typical = (high + low + close) / 3.0
    cum_tp_vol = (typical * volume).cumsum()
    cum_vol = volume.cumsum()
    return cum_tp_vol / cum_vol.replace(0, np.nan)

def obv_series(close: pd.Series, volume: pd.Series):
    sign = np.sign(close.diff().fillna(0))
    return (sign * volume).fillna(0).cumsum()

def psy_series(close: pd.Series, length=12):
    up = (close.diff() > 0).astype(int)
    return up.rolling(length, min_periods=length).mean() * 100

def williams_r(high: pd.Series, low: pd.Series, close: pd.Series, length=14):
    hh = high.rolling(length, min_periods=length).max()
    ll = low.rolling(length, min_periods=length).min()
    return -100 * (hh - close) / (hh - ll)

def bias_series(close: pd.Series, length=6):
    ma = close.rolling(length, min_periods=length).mean()
    return (close - ma) / ma * 100

# ---------- 欄位正規化 & 資料抓取 ----------
def normalize_ohlcv_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    if isinstance(df.columns, pd.MultiIndex):
        # 試著把 OHLCV 放在最外層
        fields = {"open","high","low","close","adj close","volume"}
        lvl0 = [str(x).strip().lower() for x in df.columns.get_level_values(0)]
        lvl1 = [str(x).strip().lower() for x in df.columns.get_level_values(1)]
        if set(lvl0) & fields:
            df.columns = df.columns.get_level_values(0)
        elif set(lvl1) & fields:
            df.columns = df.columns.get_level_values(1)
        else:
            df.columns = [str(a) for a, *_ in df.columns]

    mapping = {
        "open": "Open", "high": "High", "low": "Low",
        "close": "Close", "adjclose": "Adj Close", "adj close": "Adj Close",
        "adjusted close": "Adj Close", "volume": "Volume", "price": "Close",
    }
    new_cols = []
    for c in df.columns:
        key = str(c).strip().lower().replace("_", " ").replace("-", " ")
        new_cols.append(mapping.get(key, c))
    df.columns = new_cols

    # 若仍沒有 Close/Adj Close，嘗試模糊抓
    if "Close" not in df.columns and "Adj Close" not in df.columns:
        for c in list(df.columns):
            if "close" in str(c).lower():
                df.rename(columns={c: "Close"}, inplace=True)
                break
    return df

@st.cache_data(show_spinner=False)
def yf_download(ticker: str, start: dt.date, end: dt.date) -> pd.DataFrame:
    df = yf.download(
        ticker,
        start=start,
        end=end + dt.timedelta(days=1),  # 包含結束日
        auto_adjust=False,
        progress=False,
        threads=False,
        group_by="column",
    )
    if df is None or df.empty:
        return pd.DataFrame()
    df = normalize_ohlcv_columns(df)
    return df.dropna(how="all")

def add_indicators(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = normalize_ohlcv_columns(df.copy())

    # 選價格欄
    if "Adj Close" in out.columns:
        price = out["Adj Close"]; price_name = "Adj Close"
    elif "Close" in out.columns:
        price = out["Close"]; price_name = "Close"
    else:
        if all(c in out.columns for c in ["Open","High","Low"]):
            price = (out["Open"] + out["High"] + out["Low"]) / 3.0
            out["Close"] = price
            price_name = "Close*synthetic"
        else:
            return pd.DataFrame()

    for col in ["Open","High","Low","Close","Volume"]:
        if col not in out.columns:
            out[col] = price if col != "Volume" else 0

    # 指標
    out["MA3"]  = price.rolling(3,  min_periods=3).mean()
    out["MA5"]  = price.rolling(5,  min_periods=5).mean()
    out["MA10"] = price.rolling(10, min_periods=10).mean()
    out["VWAP"] = vwap_series(out["High"], out["Low"], price, out["Volume"])

    macd, signal, hist = macd_series(price, 12, 26, 9)
    out["MACD"], out["SIGNAL"], out["HIST"] = macd, signal, hist

    plus_di, minus_di, adx = dmi_adx(out["High"], out["Low"], price, 14)
    out["+DI"], out["-DI"], out["ADX"] = plus_di, minus_di, adx

    out["RSI"] = rsi_series(price, 14)
    k, d, j = stoch_kdj(out["High"], out["Low"], price, 14, 3, 3)
    out["%K"], out["%D"], out["%J"] = k, d, j

    out["OBV"]   = obv_series(price, out["Volume"])
    out["PSY"]   = psy_series(price, 12)
    out["W%R"]   = williams_r(out["High"], out["Low"], price, 14)
    out["BIAS6"] = bias_series(price, 6)

    out.attrs["price_name"] = price_name
    return out

def pick_price_column(df: pd.DataFrame) -> str | None:
    for c in ["Close","Adj Close","Price"]:
        if c in df.columns:
            return c
    for c in df.columns:
        if "close" in str(c).lower():
            return c
    return None

def plot_six_panel(ticker: str, df: pd.DataFrame):
    """回傳 matplotlib Figure（在 Streamlit 用 st.pyplot 顯示）"""
    if df is None or df.empty:
        return None

    price_col = pick_price_column(df)
    if price_col is None:
        return None

    fig, axes = plt.subplots(6, 1, sharex=True, figsize=(12, 9))
    ax1, ax2, ax3, ax4, ax5, ax6 = axes

    # 1) Price + MAs + VWAP + Volume
    ax1.set_title(f"{ticker} - Price/MA/VWAP", loc="left")
    ax1.plot(df.index, df[price_col], label=price_col)
    for col in ["MA3","MA5","MA10","VWAP"]:
        if col in df.columns:
            ax1.plot(df.index, df[col], label=col)
    ax1.grid(True, alpha=0.3)
    ax1.legend(loc="upper left", ncol=5, fontsize=8)
    ax1b = ax1.twinx()
    if "Volume" in df.columns:
        ax1b.bar(df.index, df["Volume"], alpha=0.25, width=1.0, label="Volume")
        ax1b.set_ylabel("Volume")

    # 2) MACD
    ax2.set_title("MACD", loc="left")
    if {"MACD","SIGNAL","HIST"}.issubset(df.columns):
        ax2.plot(df.index, df["MACD"], label="MACD")
        ax2.plot(df.index, df["SIGNAL"], label="Signal")
        ax2.bar(df.index, df["HIST"], width=1.0, alpha=0.4, label="Hist")
    ax2.grid(True, alpha=0.3)
    ax2.legend(loc="upper left", ncol=3, fontsize=8)

    # 3) DMI
    ax3.set_title("DMI", loc="left")
    for col in ["+DI","-DI","ADX"]:
        if col in df.columns:
            ax3.plot(df.index, df[col], label=col)
    ax3.grid(True, alpha=0.3)
    ax3.legend(loc="upper left", ncol=3, fontsize=8)

    # 4) RSI
    ax4.set_title("RSI(14)", loc="left")
    if "RSI" in df.columns:
        ax4.plot(df.index, df["RSI"], label="RSI")
        ax4.axhline(70, ls="--", lw=1)
        ax4.axhline(30, ls="--", lw=1)
    ax4.grid(True, alpha=0.3)
    ax4.legend(loc="upper left", fontsize=8)

    # 5) KD
    ax5.set_title("KD", loc="left")
    for col in ["%K","%D","%J"]:
        if col in df.columns:
            ax5.plot(df.index, df[col], label=col)
    ax5.grid(True, alpha=0.3)
    ax5.legend(loc="upper left", ncol=3, fontsize=8)

    # 6) OBV / PSY / W%R / BIAS6
    ax6.set_title("OBV / PSY / W%R / BIAS6", loc="left")
    for col in ["OBV","PSY","W%R","BIAS6"]:
        if col in df.columns:
            ax6.plot(df.index, df[col], label=col)
    ax6.grid(True, alpha=0.3)
    ax6.legend(loc="upper left", ncol=4, fontsize=8)

    plt.tight_layout()
    return fig

def build_excel_report(data_map: dict[str, pd.DataFrame]) -> bytes:
    """把每檔圖 + 數據寫到一份 Excel，回傳檔案 bytes（給 download_button）"""
    from openpyxl import Workbook
    wb = Workbook()
    # 改名預設工作表
    ws0 = wb.active
    ws0.title = "README"
    ws0["A1"] = f"Generated at {pd.Timestamp.now()}"

    for ticker, df in data_map.items():
        # 畫圖 → 轉成圖片物件
        fig = plot_six_panel(ticker, df)
        if fig is None:
            continue
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)

        # 圖表分頁
        ws_chart = wb.create_sheet(title=ticker[:31])
        img = XLImage(buf)
        img.anchor = "A1"
        ws_chart.add_image(img)

        # 數據分頁
        ws_data = wb.create_sheet(title=(ticker + "_data")[:31])
        cols = ["Date"] + list(df.columns)
        ws_data.append(cols)
        for _, row in df.reset_index().iterrows():
            values = [row["Date"]]
            for c in df.columns:
                v = row[c]
                values.append(None if pd.isna(v) else (float(v) if isinstance(v, (np.floating, float, np.integer, int)) else v))
            ws_data.append(values)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# ---------- Streamlit UI ----------
st.set_page_config(page_title="TA 6-Panel Dashboard", layout="wide")

st.title("📈 台股技術分析 6 子圖 Dashboard")
st.caption("選擇日期區間與股票，會即時抓取 Yahoo Finance 資料並繪圖。")

with st.sidebar:
    st.header("設定")
    # 日期區間
    today = dt.date.today()
    default_start = today - dt.timedelta(days=180)
    start_date = st.date_input("開始日期", value=default_start, max_value=today)
    end_date = st.date_input("結束日期", value=today, min_value=start_date, max_value=today)

    # 來源：上傳 excel 或手動輸入
    src = st.radio("股票清單來源", ["上傳 tickers.xlsx", "手動輸入"], index=0)
    tickers = []
    if src == "上傳 tickers.xlsx":
        up = st.file_uploader("上傳 Excel（第一欄名為 Ticker）", type=["xlsx"])
        if up is not None:
            try:
                df_codes = pd.read_excel(up)
                col = [c for c in df_codes.columns if str(c).strip().lower() == "ticker"]
                if col:
                    tickers = df_codes[col[0]].dropna().astype(str).tolist()
            except Exception as e:
                st.warning(f"讀取 Excel 失敗：{e}")
    else:
        tickers_text = st.text_area("輸入代碼（逗號或換行分隔）", "2330.TW, 2317.TW, 2454.TW")
        tickers = [t.strip() for t in tickers_text.replace("\n", ",").split(",") if t.strip()]

    tickers = list(dict.fromkeys(tickers))  # 去重
    go = st.button("產生圖表 / 下載報告")

# 主區域
if go:
    if not tickers:
        st.error("請提供至少一個代碼")
        st.stop()

    tab_charts, tab_table, tab_export = st.tabs(["📊 圖表", "📄 最後一日數值", "⬇️ 匯出 Excel"])

    all_data: dict[str, pd.DataFrame] = {}
    last_rows = []

    with st.spinner("抓取資料與繪圖中…"):
        for t in tickers:
            raw = yf_download(t, start_date, end_date)
            if raw.empty:
                st.warning(f"[{t}] 無資料，略過")
                continue
            df = add_indicators(raw)
            if df.empty:
                st.warning(f"[{t}] 無法計算指標，略過")
                continue

            all_data[t] = df

            # 圖表
            with tab_charts:
                st.subheader(t)
                fig = plot_six_panel(t, df)
                if fig is None:
                    st.warning(f"[{t}] 無可用價欄，無法繪圖。")
                else:
                    st.pyplot(fig, clear_figure=True)

            # 最後一日數值
            last = df.iloc[-1]
            last_rows.append({
                "Ticker": t,
                "Date": df.index[-1].date(),
                "Close": float(last.get("Close", np.nan)) if pd.notna(last.get("Close", np.nan)) else np.nan,
                "MA5": float(last.get("MA5", np.nan)) if pd.notna(last.get("MA5", np.nan)) else np.nan,
                "RSI": float(last.get("RSI", np.nan)) if pd.notna(last.get("RSI", np.nan)) else np.nan,
                "MACD": float(last.get("MACD", np.nan)) if pd.notna(last.get("MACD", np.nan)) else np.nan,
                "SIGNAL": float(last.get("SIGNAL", np.nan)) if pd.notna(last.get("SIGNAL", np.nan)) else np.nan,
                "ADX": float(last.get("ADX", np.nan)) if pd.notna(last.get("ADX", np.nan)) else np.nan,
                "+DI": float(last.get("+DI", np.nan)) if pd.notna(last.get("+DI", np.nan)) else np.nan,
                "-DI": float(last.get("-DI", np.nan)) if pd.notna(last.get("-DI", np.nan)) else np.nan,
                "%K": float(last.get("%K", np.nan)) if pd.notna(last.get("%K", np.nan)) else np.nan,
                "%D": float(last.get("%D", np.nan)) if pd.notna(last.get("%D", np.nan)) else np.nan,
                "%J": float(last.get("%J", np.nan)) if pd.notna(last.get("%J", np.nan)) else np.nan,
                "OBV": float(last.get("OBV", np.nan)) if pd.notna(last.get("OBV", np.nan)) else np.nan,
                "PSY": float(last.get("PSY", np.nan)) if pd.notna(last.get("PSY", np.nan)) else np.nan,
                "W%R": float(last.get("W%R", np.nan)) if pd.notna(last.get("W%R", np.nan)) else np.nan,
                "BIAS6": float(last.get("BIAS6", np.nan)) if pd.notna(last.get("BIAS6", np.nan)) else np.nan,
            })

    # 彙總表格
    if last_rows:
        with tab_table:
            st.dataframe(pd.DataFrame(last_rows), use_container_width=True)

    # 匯出 Excel
    if all_data:
        with tab_export:
            xls_bytes = build_excel_report(all_data)
            st.download_button(
                label="下載 Excel 報告",
                data=xls_bytes,
                file_name="ETF100_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
